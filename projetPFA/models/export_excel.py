"""
models/export_excel.py
------------------------
Bloc 5 - GENERATION DU NOUVEAU FICHIER EXCEL.

Classe ExportExcel : génère et enregistre le fichier Excel propre
(Donnees_Nettoyees.xlsx) — lignes utiles uniquement, colonnes pertinentes
conservées — ainsi qu'un rapport d'audit des lignes rejetées. Ce fichier
alimente ensuite Power BI (hors code, étape 6 du diagramme de séquence).
"""

from __future__ import annotations

import io
import json
import os
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl.drawing.image import Image as OpenpyxlImage

from config import (
    COLONNE_CIBLE,
    DEFAULT_AUDIT_NAME,
    DEFAULT_OUTPUT_NAME,
    FEUILLE_DONNEES_VALIDES,
    FEUILLE_LIGNES_REJETEES,
    OUTPUT_DIR,
    SEUIL_COLONNE_VIDE_EXPORT,
)
from models.logger import get_logger
from models.modele_ml import CLASSES_UTILITE, generer_labels_utilite

logger = get_logger(__name__)


@dataclass
class ExportExcel:
    cheminSortie: str = OUTPUT_DIR

    # ------------------------------------------------------------------
    def colonnes_pertinentes(
        self, df: pd.DataFrame, seuil_vide: float = SEUIL_COLONNE_VIDE_EXPORT
    ) -> list[str]:
        """Sélectionne, de façon GENERIQUE, les colonnes "pertinentes" à
        conserver dans le fichier final : celles qui ne sont pas quasiment
        entièrement vides (au-delà de `seuil_vide` % de valeurs manquantes),
        ni constantes (une seule valeur distincte sur tout le fichier)."""
        colonnes = []
        n = len(df)
        for col in df.columns:
            if n and df[col].isna().mean() > seuil_vide:
                continue
            if df[col].nunique(dropna=True) <= 1:
                continue
            colonnes.append(col)
        return colonnes or list(df.columns)

    def genererExcel(self, df: pd.DataFrame, nom_feuille: str = "Donnees") -> bytes:
        """Génère un fichier Excel en mémoire (bytes), prêt pour un bouton de téléchargement."""
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=nom_feuille)
        return buffer.getvalue()

    def enregistrer(
        self, df: pd.DataFrame, nom_fichier: str = DEFAULT_OUTPUT_NAME, nom_feuille: str = "Donnees"
    ) -> str:
        """Enregistre le fichier Excel sur le disque, dans le dossier outputs/."""
        os.makedirs(self.cheminSortie, exist_ok=True)
        chemin_complet = os.path.join(self.cheminSortie, nom_fichier)
        df.to_excel(chemin_complet, index=False, sheet_name=nom_feuille)
        logger.info("Fichier exporté : %s (%d lignes)", chemin_complet, len(df))
        return chemin_complet

    def genererRapportAudit(self, df_rejete: pd.DataFrame) -> bytes:
        return self.genererExcel(df_rejete, nom_feuille=FEUILLE_LIGNES_REJETEES)

    def _filtrer_df_utiles(self, df: pd.DataFrame) -> pd.DataFrame:
        """Retourne le DataFrame nettoyé en ne gardant que les lignes utiles."""
        df = df.dropna(how="all").dropna(axis=1, how="all")
        if df.empty:
            return df

        if COLONNE_CIBLE in df.columns:
            colonne = df[COLONNE_CIBLE].astype(str).str.strip()
            if colonne.isin(CLASSES_UTILITE).any():
                mask = colonne == CLASSES_UTILITE[1]
            else:
                mask = pd.to_numeric(colonne, errors="coerce") == 1
            return df.loc[mask].reset_index(drop=True)

        labels, _ = generer_labels_utilite(df)
        return df.loc[labels.astype(bool)].reset_index(drop=True)

    def genererExportNettoyeDepuisFichier(self, chemin_source: str) -> bytes:
        """Génère un fichier Excel nettoyé en préservant la structure d'origine."""
        buffer = io.BytesIO()
        if chemin_source.lower().endswith(".csv"):
            df = pd.read_csv(chemin_source)
            df_nettoye = self._filtrer_df_utiles(df)
            nom_feuille = Path(chemin_source).stem[:31]
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                df_nettoye.to_excel(writer, index=False, sheet_name=nom_feuille)
            return buffer.getvalue()

        xls = pd.ExcelFile(chemin_source)
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            for feuille in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=feuille)
                df_nettoye = self._filtrer_df_utiles(df)
                df_nettoye.to_excel(writer, index=False, sheet_name=feuille[:31])
        return buffer.getvalue()

    # ------------------------------------------------------------------
    # Export générique "2 onglets" : lignes utiles (colonnes pertinentes
    # uniquement) + lignes rejetées (audit, toutes colonnes conservées).
    # Remplace l'export "multi-KPI" métier de la version précédente.
    # ------------------------------------------------------------------
    def genererExportComplet(
        self, df_utile: pd.DataFrame, df_rejete: pd.DataFrame
    ) -> bytes:
        colonnes_ok = self.colonnes_pertinentes(df_utile)
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df_utile[colonnes_ok].to_excel(
                writer, index=False, sheet_name=FEUILLE_DONNEES_VALIDES[:31]
            )
        return buffer.getvalue()

    def genererExportAvecGraphiques(
        self,
        df_utile: pd.DataFrame,
        df_rejete: pd.DataFrame,
        analyses: dict[str, pd.DataFrame] | None = None,
        figures: dict[str, plt.Figure] | None = None,
        nom_fichier: str = DEFAULT_OUTPUT_NAME,
    ) -> bytes:
        """Génère un fichier Excel contenant :
        - onglet des données utiles (colonnes pertinentes),
        - onglet des lignes rejetées (audit),
        - un onglet par table d'analyse fournie,
        - un onglet "Graphiques" avec les figures fournies insérées
        en tant qu'images PNG.

        `analyses` : dict nom->DataFrame
        `figures` : dict nom->matplotlib.figure.Figure
        """
        analyses = analyses or {}
        figures = figures or {}

        colonnes_ok = self.colonnes_pertinentes(df_utile)
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            # Données utiles
            df_utile[colonnes_ok].to_excel(
                writer, index=False, sheet_name=FEUILLE_DONNEES_VALIDES[:31]
            )

            # Lignes rejetées (audit)
            if df_rejete is not None and not df_rejete.empty:
                df_rejete.to_excel(
                    writer, index=False, sheet_name=FEUILLE_LIGNES_REJETEES[:31]
                )

            # Analyses tabulaires
            for nom, df_an in analyses.items():
                safe_name = str(nom)[:31]
                try:
                    df_an.to_excel(writer, index=True, sheet_name=safe_name)
                except Exception:
                    # si nom du sheet déjà existant ou autre erreur, suffixer
                    df_an.to_excel(writer, index=True, sheet_name=(safe_name + "_1")[:31])

            # Insérer les images dans un onglet Graphiques (openpyxl)
            wb = writer.book
            if figures:
                if "Graphiques" in wb.sheetnames:
                    ws = wb["Graphiques"]
                else:
                    ws = wb.create_sheet(title="Graphiques")

                row = 1
                for titre, fig in figures.items():
                    # Sauvegarde figure en PNG en mémoire
                    img_buf = io.BytesIO()
                    fig.savefig(img_buf, bbox_inches="tight", dpi=150)
                    img_buf.seek(0)
                    img = OpenpyxlImage(img_buf)

                    # Positionner l'image à la ligne courante
                    cell_ref = f"A{row}"
                    ws.add_image(img, cell_ref)

                    # Avancer la ligne : approx hauteur image en lignes
                    row += int(img.height / 20) + 4

        return buffer.getvalue()

    def _safe_sheet_name(self, name: str) -> str:
        return str(name).replace("/", "_").replace("\\", "_")[:31]

    def _ecrire_df_sur_feuille(
        self, writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str, index: bool = False
    ) -> None:
        df.to_excel(writer, index=index, sheet_name=self._safe_sheet_name(sheet_name))

    def genererRapportStructureExcel(
        self,
        df_utile: pd.DataFrame | None = None,
        df_rejete: pd.DataFrame | None = None,
        analyses: dict[str, pd.DataFrame] | None = None,
        meta: dict[str, str] | None = None,
        figures: dict[str, plt.Figure] | None = None,
        nom_fichier: str = DEFAULT_OUTPUT_NAME,
    ) -> bytes:
        """Génère un rapport structuré réutilisable dans un autre outil.

        Le classeur Excel contient des feuilles séparées pour :
        - métadonnées
        - statistiques descriptives
        - valeurs manquantes
        - agrégations par catégorie
        - données utiles / lignes rejetées
        - graphiques facultatifs
        """
        analyses = analyses or {}
        meta = meta or {}
        figures = figures or {}

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            if meta:
                pd.DataFrame(list(meta.items()), columns=["cle", "valeur"]).to_excel(
                    writer, index=False, sheet_name="Meta"
                )

            if df_utile is not None:
                colonnes_ok = self.colonnes_pertinentes(df_utile)
                if colonnes_ok:
                    self._ecrire_df_sur_feuille(
                        writer,
                        df_utile[colonnes_ok],
                        FEUILLE_DONNEES_VALIDES,
                        index=False,
                    )
                else:
                    self._ecrire_df_sur_feuille(writer, df_utile, FEUILLE_DONNEES_VALIDES, index=False)

            if df_rejete is not None and not df_rejete.empty:
                self._ecrire_df_sur_feuille(
                    writer, df_rejete, FEUILLE_LIGNES_REJETEES, index=False
                )

            for nom, df_an in analyses.items():
                self._ecrire_df_sur_feuille(writer, df_an.reset_index(), nom, index=False)

            if figures:
                wb = writer.book
                ws = wb.create_sheet(title=self._safe_sheet_name("Graphiques"))
                row = 1
                for titre, fig in figures.items():
                    img_buf = io.BytesIO()
                    fig.savefig(img_buf, bbox_inches="tight", dpi=150)
                    img_buf.seek(0)
                    img = OpenpyxlImage(img_buf)
                    ws.add_image(img, f"A{row}")
                    row += int(img.height / 20) + 4

        return buffer.getvalue()

    def genererRapportJSON(
        self,
        analyses: dict[str, pd.DataFrame] | None = None,
        meta: dict[str, str] | None = None,
        df_utile: pd.DataFrame | None = None,
        df_rejete: pd.DataFrame | None = None,
    ) -> bytes:
        """Génère un rapport structuré sérialisable en JSON pour réutilisation."""
        analyses = analyses or {}
        meta = meta or {}

        report: dict[str, object] = {
            "meta": meta,
            "analyses": {},
        }

        for nom, df_an in analyses.items():
            report["analyses"][str(nom)] = df_an.fillna("").astype(str).to_dict(orient="records")

        if df_utile is not None:
            report["donnees_utiles"] = df_utile.fillna("").astype(str).to_dict(orient="records")

        if df_rejete is not None:
            report["lignes_rejetees"] = df_rejete.fillna("").astype(str).to_dict(orient="records")

        return json.dumps(report, indent=2, ensure_ascii=False).encode("utf-8")

    def _ajouter_pied_de_page(self, fig: plt.Figure, page_num: int) -> None:
        """Ajoute un pied de page numéroté en bas de la page."""
        fig.text(
            0.5,
            0.02,
            f"Page {page_num}",
            ha="center",
            va="bottom",
            fontsize=9,
            color="#555555",
        )

    def _figure_a_un_titre(self, fig: plt.Figure) -> bool:
        """Détecte si une figure contient déjà un titre ou un texte d'en-tête."""
        if getattr(fig, "_suptitle", None) is not None:
            text = fig._suptitle.get_text().strip()
            if text:
                return True

        for ax in fig.axes:
            if ax.get_title().strip():
                return True

        for texte in fig.texts:
            if texte.get_text().strip() and texte.get_position()[1] > 0.8:
                return True

        return False

    def genererRapportPDF(
        self,
        analyses: dict[str, pd.DataFrame] | None = None,
        figures: dict[str, plt.Figure] | None = None,
        meta: dict[str, str] | None = None,
        titre: str = "Rapport d'analyse détaillé",
    ) -> bytes:
        """Génère un rapport PDF contenant les analyses tabulaires et les figures."""
        analyses = analyses or {}
        figures = figures or {}
        meta = meta or {}

        buffer = io.BytesIO()
        page_num = 1
        with PdfPages(buffer) as pdf:
            # Page de garde
            fig = plt.figure(figsize=(11.69, 8.27), facecolor="#FFFFFF")
            fig.suptitle(titre, fontsize=28, y=0.92, weight="bold")
            subtitle = "Rapport d'analyse détaillée du fichier nettoyé"
            fig.text(0.1, 0.82, subtitle, fontsize=14, color="#333333")

            meta_text = "\n".join(f"{cle} : {val}" for cle, val in meta.items())
            if not meta_text:
                meta_text = "Rapport d'analyse généré automatiquement."

            fig.text(
                0.1,
                0.72,
                "Résumé du traitement:",
                fontsize=12,
                weight="bold",
                color="#444444",
            )
            fig.text(
                0.1,
                0.68,
                meta_text,
                fontsize=11,
                color="#444444",
                va="top",
            )

            fig.text(
                0.1,
                0.45,
                "Ce rapport contient :",
                fontsize=12,
                weight="bold",
                color="#444444",
            )
            fig.text(
                0.1,
                0.40,
                "- Statistiques descriptives\n- Valeurs manquantes\n- Agrégations clés\n- Visualisations et graphiques utiles",
                fontsize=11,
                color="#444444",
                va="top",
            )

            fig.text(
                0.1,
                0.15,
                "Généré par le pipeline de nettoyage intelligent de données.",
                fontsize=10,
                color="#666666",
            )

            self._ajouter_pied_de_page(fig, page_num)
            pdf.savefig(fig, bbox_inches="tight")
            plt.close(fig)
            page_num += 1

            # Pages de tableaux d'analyse
            for nom, df_an in analyses.items():
                df_disp = df_an.copy()
                affiche_full = True
                if len(df_disp) > 30:
                    df_disp = df_disp.head(30)
                    affiche_full = False
                if len(df_disp.columns) > 12:
                    df_disp = df_disp.iloc[:, :12]
                    affiche_full = False

                fig, ax = plt.subplots(figsize=(11.69, 8.27), facecolor="#FFFFFF")
                ax.axis("off")
                fig.suptitle(nom, fontsize=18, y=0.95, weight="bold")

                col_count = len(df_disp.columns)
                col_width = min(0.18, max(0.05, 0.9 / max(1, col_count)))
                table = ax.table(
                    cellText=df_disp.fillna("").astype(str).values,
                    colLabels=df_disp.columns,
                    loc="center",
                    cellLoc="left",
                    colWidths=[col_width] * col_count,
                    colColours=["#F2E4D5"] * col_count,
                    cellColours=[["#FFFFFF" if i % 2 == 0 else "#FBF6F0" for _ in df_disp.columns] for i in range(len(df_disp))],
                )
                table.auto_set_font_size(False)
                table.set_fontsize(8)
                table.scale(1, 1.3)
                try:
                    table.auto_set_column_width(col=list(range(col_count)))
                except Exception:
                    pass

                note = ""
                if not affiche_full:
                    note = (
                        "Affichage limité aux 30 premières lignes et/ou 12 premières colonnes "
                        "pour conserver une présentation lisible."
                    )
                    fig.text(0.1, 0.08, note, fontsize=10, color="#555555")

                self._ajouter_pied_de_page(fig, page_num)
                pdf.savefig(fig, bbox_inches="tight")
                plt.close(fig)
                page_num += 1

            # Pages de figures
            for nom, fig in figures.items():
                fig.set_size_inches(11.69, 8.27)
                if not self._figure_a_un_titre(fig):
                    fig.suptitle(nom, fontsize=18, y=0.95, weight="bold")
                    top_rect = [0, 0.05, 1, 0.90]
                else:
                    top_rect = [0, 0.05, 1, 0.94]

                try:
                    fig.tight_layout(rect=top_rect)
                except Exception:
                    fig.tight_layout()
                self._ajouter_pied_de_page(fig, page_num)
                pdf.savefig(fig, bbox_inches="tight")
                plt.close(fig)
                page_num += 1

        buffer.seek(0)
        return buffer.getvalue()

    def enregistrerExportComplet(
        self,
        df_utile: pd.DataFrame,
        df_rejete: pd.DataFrame,
        nom_fichier: str = DEFAULT_OUTPUT_NAME,
    ) -> str:
        os.makedirs(self.cheminSortie, exist_ok=True)
        chemin_complet = os.path.join(self.cheminSortie, nom_fichier)
        colonnes_ok = self.colonnes_pertinentes(df_utile)
        with pd.ExcelWriter(chemin_complet, engine="openpyxl") as writer:
            df_utile[colonnes_ok].to_excel(
                writer, index=False, sheet_name=FEUILLE_DONNEES_VALIDES[:31]
            )
        logger.info(
            "Fichier nettoyé exporté : %s (%d lignes utiles)",
            chemin_complet, len(df_utile),
        )
        return chemin_complet
